import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def limpiar_colores(texto):
    """
    Cuando NO hay 'Detalle paquete':
    - Divide por guiones "-"
    - Se queda con lo que está antes de ":" (ignora tallas)
    - Une los colores con " - "
    """
    partes = str(texto).split("-")
    colores = []
    for parte in partes:
        parte = parte.strip()
        if not parte:
            continue
        if ":" in parte:
            colores.append(parte.split(":")[0].strip())  # lo que está antes de ":"
        else:
            colores.append(parte)
    return " - ".join(colores)


def construir_color(row):
    lista_detalle = str(row["Lista de detalle color y talla"])
    detalle_paquete = str(row.get("Detalle paquete", "")).strip()

    # Si no hay detalle_paquete -> usamos limpiar_colores
    if not detalle_paquete or detalle_paquete.lower() == "nan":
        valor = limpiar_colores(lista_detalle)
        return "" if valor.lower() == "nan" else valor

    # Capturar códigos: Sxx o números de 3 dígitos
    codigos = re.findall(r'(S\d+|[A-Z]\d{2,3}|\d{3})', lista_detalle, flags=re.IGNORECASE)

    # Caso: SOLO UN código → se pega todo el detalle completo
    if len(codigos) == 1:
        return f"{codigos[0]} {detalle_paquete}"

    # Caso: Varios códigos → separar detalle SOLO por "-"
    colores_paquete = [c.strip() for c in detalle_paquete.split("-") if c.strip()]

    # Emparejar códigos con descripciones del paquete
    resultado = []
    for i, codigo in enumerate(codigos):
        if i < len(colores_paquete):
            resultado.append(f"{codigo} {colores_paquete[i]}")
        else:
            resultado.append(codigo)  # Si faltan colores, dejamos solo el código

    return " - ".join(resultado)


def actualizar_photolist(archivo_origen, archivo_destino, archivo_paginas, nombre_disenador):
    df_origen = pd.read_excel(archivo_origen, sheet_name="Por Veh-Pág")

    if 'Lista de detalle color y talla' not in df_origen.columns:
        raise ValueError("La columna 'Lista de detalle color y talla' no está en el archivo.")

    if 'ClaseVenta' not in df_origen.columns:
        raise ValueError("La columna 'ClaseVenta' no está en el archivo.")

    # Construcción de colores (usa Detalle paquete si existe)
    df_origen['Color y talla'] = df_origen.apply(construir_color, axis=1)

    paginas_validas_df = pd.read_excel(archivo_paginas)
    if 'Paginas' not in paginas_validas_df.columns or 'Vehículo' not in paginas_validas_df.columns:
        raise ValueError("El archivo 'Paginas CR.xlsx' debe tener las columnas 'Paginas' y 'Vehículo'.")

    paginas_validas_df['Paginas'] = paginas_validas_df['Paginas'].astype(str)
    paginas_validas_df['Vehículo'] = paginas_validas_df['Vehículo'].astype(str)

    df_origen['Página'] = df_origen['Página'].astype(str)
    df_origen['Vehículo'] = df_origen['Vehículo'].astype(str)

    combinaciones_validas = set(zip(paginas_validas_df['Paginas'], paginas_validas_df['Vehículo']))
    df_origen = df_origen[df_origen.apply(lambda row: (row['Página'], row['Vehículo']) in combinaciones_validas, axis=1)]

    columnas_finales = ['Pais', 'DISEÑADOR', 'Página', 'Referencia', 'Color y talla']
    filas_expand = []
    filas_por_grupo = []
    filas_separadoras = []  # Guardaremos las filas vacías para colorearlas

    # Limpiar espacios extra y guiones al final
    df_origen['Color y talla'] = df_origen['Color y talla'].str.strip()

    def join_colores_sin_guion_final(colores):
        colores = [c.strip() for c in colores if c.strip()]
        return ' - '.join(colores)

    # Agrupar y unir colores sin generar guion final innecesario
    df_agrupado = (
        df_origen
        .groupby(['Pais', 'Referencia', 'Página'])['Color y talla']
        .apply(lambda x: join_colores_sin_guion_final(x))
        .reset_index()
    )

    # Ordenar
    df_agrupado = df_agrupado.sort_values(['Pais', 'Página', 'Referencia']).reset_index(drop=True)

    grupos = df_agrupado.groupby(['Pais', 'Página'])

    for (pais, pagina), grupo in grupos:
        start_idx = len(filas_expand)

        # Filas reales
        for idx, fila in grupo.iterrows():
            filas_expand.append({
                'Pais': pais,
                'DISEÑADOR': nombre_disenador,
                'Página': pagina,
                'Referencia': fila['Referencia'],
                'Color y talla': fila['Color y talla']
            })

        # Rellenar hasta 14 filas
        n_filas_actual = len(grupo)
        while n_filas_actual < 14:
            filas_expand.append({
                'Pais': pais,
                'DISEÑADOR': nombre_disenador,
                'Página': pagina,
                'Referencia': '',
                'Color y talla': ''
            })
            n_filas_actual += 1

        end_idx = len(filas_expand)
        filas_por_grupo.append((start_idx + 2, end_idx + 1))

        # Fila separadora vacía
        filas_expand.append({
            'Pais': '',
            'DISEÑADOR': '',
            'Página': '',
            'Referencia': '',
            'Color y talla': ''
        })
        filas_separadoras.append(len(filas_expand))  # Guardamos índice para pintar

    # Crear DataFrame final
    df_final = pd.DataFrame(filas_expand)
    df_final = df_final[columnas_finales]
    df_final.to_excel(archivo_destino, index=False)

    wb = load_workbook(archivo_destino)
    ws = wb.active

    # Pintar filas separadoras de gris oscuro
    gray_fill = PatternFill(start_color="BCBCBC", end_color="BCBCBC", fill_type="solid")
    for row_idx in filas_separadoras:
        for col in range(1, 6):  # Columnas A-E
            ws.cell(row=row_idx + 1, column=col).fill = gray_fill

    # Merge columnas A-C
    columnas_a_mergear = ['A', 'B', 'C']
    for col in columnas_a_mergear:
        for start, end in filas_por_grupo:
            if start < end:
                ws.merge_cells(f"{col}{start}:{col}{end}")


    wb.save(archivo_destino)
    return df_final


# -------------------- Streamlit UI --------------------
st.title("Actualizador de PHOTOLIST")
nombre_disenador = st.text_input("👤 Escribe tu nombre (Diseñador)")
archivo_origen = st.file_uploader("Sube el archivo de origen (Por Veh-Pág)", type=["xlsx"])
archivo_paginas = st.file_uploader("Sube el archivo de páginas válidas (Paginas.xlsx)", type=["xlsx"])

if archivo_origen and archivo_paginas and nombre_disenador and st.button("Procesar"):
    try:
        nombre_archivo_salida = f"PHOTOLIST_{nombre_disenador.replace(' ', '_')}.xlsx"
        df_resultado = actualizar_photolist(archivo_origen, nombre_archivo_salida, archivo_paginas, nombre_disenador)

        st.success(f"Archivo generado exitosamente: {nombre_archivo_salida}")
        st.dataframe(df_resultado)

        with open(nombre_archivo_salida, "rb") as f:
            bytes_data = f.read()

        st.download_button(
            label="📥 Descargar archivo Excel",
            data=bytes_data,
            file_name=nombre_archivo_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Hubo un error: {e}")
elif not nombre_disenador:
    st.warning("⚠️ Por favor, ingresa tu nombre antes de procesar.")
